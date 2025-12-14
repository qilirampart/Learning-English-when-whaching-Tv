"""
导出服务 - 单词本导出为 Excel 和 PDF
"""

from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ExportService:
    """导出服务类"""

    @staticmethod
    def export_to_excel(words, user_info=None):
        """
        导出单词到 Excel

        Args:
            words: 单词列表
            user_info: 用户信息（可选）

        Returns:
            BytesIO: Excel 文件的字节流
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 库未安装，请运行: pip install openpyxl")

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "我的单词本"

        # 设置列宽
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20

        # 标题样式
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')

        # 表头样式
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 添加标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_text = f"📚 美剧单词学习助手 - 单词本"
        if user_info:
            title_text += f" ({user_info.get('username', '未知用户')})"
        title_cell.value = title_text
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        ws.row_dimensions[1].height = 30

        # 添加导出时间
        ws.merge_cells('A2:H2')
        export_time_cell = ws['A2']
        export_time_cell.value = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        export_time_cell.alignment = Alignment(horizontal='right')
        export_time_cell.font = Font(name='Arial', size=9, italic=True)

        # 添加统计信息
        ws.merge_cells('A3:H3')
        stats_cell = ws['A3']
        stats_cell.value = f"总单词数: {len(words)}"
        stats_cell.alignment = Alignment(horizontal='center')
        stats_cell.font = Font(name='Arial', size=10, bold=True)
        ws.row_dimensions[3].height = 20

        # 添加表头
        headers = ['序号', '单词', '音标', '中文释义', '英文释义', '掌握度', '查询次数', '最后查询']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[4].height = 25

        # 数据样式
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        data_alignment_center = Alignment(horizontal='center', vertical='center')

        # 添加数据
        for idx, word in enumerate(words, 1):
            row = idx + 4

            # 序号
            cell = ws.cell(row=row, column=1, value=idx)
            cell.alignment = data_alignment_center
            cell.border = thin_border

            # 单词
            cell = ws.cell(row=row, column=2, value=word.get('word', ''))
            cell.font = Font(name='Arial', size=12, bold=True)
            cell.alignment = data_alignment
            cell.border = thin_border

            # 音标
            cell = ws.cell(row=row, column=3, value=word.get('phonetic', ''))
            cell.alignment = data_alignment
            cell.border = thin_border

            # 中文释义
            cell = ws.cell(row=row, column=4, value=word.get('translation', ''))
            cell.alignment = data_alignment
            cell.border = thin_border

            # 英文释义
            definition = word.get('definition', '')
            if len(definition) > 100:
                definition = definition[:100] + '...'
            cell = ws.cell(row=row, column=5, value=definition)
            cell.alignment = data_alignment
            cell.border = thin_border

            # 掌握度
            mastery_level = word.get('mastery_level', 0)
            cell = ws.cell(row=row, column=6, value=f"{mastery_level}/5")
            cell.alignment = data_alignment_center
            cell.border = thin_border
            # 根据掌握度设置颜色
            if mastery_level == 5:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif mastery_level >= 3:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif mastery_level > 0:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            # 查询次数
            cell = ws.cell(row=row, column=7, value=word.get('query_count', 0))
            cell.alignment = data_alignment_center
            cell.border = thin_border

            # 最后查询时间
            last_query = word.get('last_query', '')
            if last_query:
                try:
                    last_query = datetime.fromisoformat(last_query.replace('Z', '+00:00'))
                    last_query = last_query.strftime('%Y-%m-%d %H:%M')
                except:
                    pass
            cell = ws.cell(row=row, column=8, value=last_query)
            cell.alignment = data_alignment_center
            cell.border = thin_border

            # 设置行高
            ws.row_dimensions[row].height = 20

        # 冻结首行
        ws.freeze_panes = 'A5'

        # 保存到字节流
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_to_pdf(words, user_info=None):
        """
        导出单词到 PDF

        Args:
            words: 单词列表
            user_info: 用户信息（可选）

        Returns:
            BytesIO: PDF 文件的字节流
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab 库未安装，请运行: pip install reportlab")

        # 注册中文字体
        import os
        import platform

        # 尝试注册中文字体
        try:
            system = platform.system()
            if system == 'Windows':
                # Windows 系统字体路径
                font_path = 'C:/Windows/Fonts/msyh.ttc'  # 微软雅黑
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('Chinese', font_path))
                    chinese_font = 'Chinese'
                else:
                    # 尝试使用 SimHei（黑体）
                    font_path = 'C:/Windows/Fonts/simhei.ttf'
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('Chinese', font_path))
                        chinese_font = 'Chinese'
                    else:
                        chinese_font = 'Helvetica'  # 降级到默认字体
            else:
                # Linux/Mac 可以添加其他字体路径
                chinese_font = 'Helvetica'
        except:
            chinese_font = 'Helvetica'  # 如果注册失败，使用默认字体

        # 创建字节流
        output = BytesIO()

        # 创建 PDF 文档
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=60,
            bottomMargin=40
        )

        # 存储内容元素
        elements = []

        # 获取样式
        styles = getSampleStyleSheet()

        # 自定义标题样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName=chinese_font
        )

        # 自定义正文样式
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            alignment=TA_LEFT,
            fontName=chinese_font
        )

        # 标题
        title_text = "📚 美剧单词学习助手 - 单词本"
        if user_info:
            title_text += f" ({user_info.get('username', '未知用户')})"
        title = Paragraph(title_text, title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        # 导出信息
        export_info = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 总单词数: {len(words)}"
        info = Paragraph(export_info, body_style)
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))

        # 准备表格数据
        table_data = [['序号', '单词', '音标', '中文释义', '掌握度', '查询次数']]

        for idx, word in enumerate(words, 1):
            row = [
                str(idx),
                word.get('word', ''),
                word.get('phonetic', ''),
                word.get('translation', '')[:40],  # 限制长度
                f"{word.get('mastery_level', 0)}/5",
                str(word.get('query_count', 0))
            ]
            table_data.append(row)

        # 创建表格
        table = Table(table_data, colWidths=[0.6*inch, 1.2*inch, 1.0*inch, 2.5*inch, 0.8*inch, 0.8*inch])

        # 表格样式
        table_style = TableStyle([
            # 表头样式
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B9BD5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), chinese_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

            # 数据行样式
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # 序号居中
            ('ALIGN', (4, 1), (-1, -1), 'CENTER'),  # 掌握度和查询次数居中
            ('FONTNAME', (0, 1), (-1, -1), chinese_font),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # 网格线
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

            # 交替行颜色
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ])

        table.setStyle(table_style)
        elements.append(table)

        # 添加页脚信息
        elements.append(Spacer(1, 0.5*inch))
        footer = Paragraph(
            f"生成于: 美剧单词学习助手 | {datetime.now().strftime('%Y-%m-%d')}",
            ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER,
                fontName=chinese_font
            )
        )
        elements.append(footer)

        # 构建 PDF
        doc.build(elements)
        output.seek(0)

        return output

    @staticmethod
    def check_dependencies():
        """检查导出功能的依赖是否已安装"""
        return {
            'excel': OPENPYXL_AVAILABLE,
            'pdf': REPORTLAB_AVAILABLE
        }
